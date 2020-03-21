import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import time
def creat_file(name,ID):
	if(os.path.exists("{}-{}.xlsx".format(name,ID))==False):
		writer = pd.ExcelWriter('{}-{}.xlsx'.format(name,ID), engine='xlsxwriter')
		df = pd.DataFrame({'Date': [],
                   			'time_staer': [],
                   			'time_end':[]})
		df.to_excel(writer, index=False)
		writer.save()
def start(name,Id,Time=time.localtime()):
	creat_file(name,Id)
	wb= load_workbook("{}-{}.xlsx".format(name,Id))
	sheet=wb['Sheet1']
	sheet.append({'A':"{}.{}.{}".format(Time.tm_mday,Time.tm_mon,Time.tm_year) ,
		           'b':"{}:{}".format(Time.tm_hour,Time.tm_min)
		              })
	wb.save("{}-{}.xlsx".format(name,Id))
def end(name,Id):
	creat_file(name,Id)
	TT=time.localtime()
	wb= load_workbook("{}-{}.xlsx".format(name,Id))
	sheet=wb['Sheet1']
	sheet.append({'C':"{}:{}".format(TT.tm_hour,TT.tm_min)
		              })
	wb.save("{}-{}.xlsx".format(name,Id))





